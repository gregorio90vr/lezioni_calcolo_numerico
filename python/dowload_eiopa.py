"""
Scarica le curve EIOPA RFR (spot no VA, curva Euro) e produce un .xlsx.

ISTRUZIONI:
  1. Crea cartella 'eiopa_zips/' nella stessa directory di questo script
  2. Scarica i zip da EIOPA e mettili in 'eiopa_zips/'
     I nomi accettati sono due formati:
       - EIOPA_RFR_20220331.zip  (con data YYYYMMDD nel nome)
       - March 2022.zip          (con nome del mese in inglese/italiano e anno)
  3. python download_eiopa_curves.py

Comandi extra:
    python download_eiopa_curves.py missing   → lista zip mancanti
    python download_eiopa_curves.py diagnose 2022-03-31

Requisiti:
    pip install requests openpyxl pandas
"""

import io, re, sys, time, zipfile, calendar
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
START_DATE  = "2016-01-01"
END_DATE    = "2024-12-31"
OUTPUT_FILE = "EIOPA_RFR_EUR_curves.xlsx"
ZIP_FOLDER  = "eiopa_zips"
SLEEP_SEC   = 1.5
TENORS      = list(range(1, 21))
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36",
}

AUTO_DOWNLOAD_TEMPLATES = [
    "https://register.eiopa.europa.eu/Publications/Standards/EIOPA_RFR_{date}.zip",
    "https://register.eiopa.europa.eu/CEIOPS-Archive/Documents/RFR/EIOPA_RFR_{date}.zip",
]

SHEET_ALIASES = ["rfr_spot_no_va", "rfr spot no va", "spot_no_va", "spot no va"]

# Mappa nome mese → numero (inglese + italiano)
MONTH_NAMES: dict[str, int] = {
    "january": 1,  "february": 2,  "march": 3,     "april": 4,
    "may": 5,      "june": 6,      "july": 7,       "august": 8,
    "september": 9,"october": 10,  "november": 11,  "december": 12,
    "gennaio": 1,  "febbraio": 2,  "marzo": 3,      "aprile": 4,
    "maggio": 5,   "giugno": 6,    "luglio": 7,     "agosto": 8,
    "settembre": 9,"ottobre": 10,  "novembre": 11,  "dicembre": 12,
}


def _last_day(year: int, month: int) -> date:
    return date(year, month, calendar.monthrange(year, month)[1])


def _parse_month_name_zip(path: Path) -> date | None:
    """
    Prova a estrarre anno e mese da un nome tipo 'March 2022.zip'
    o 'December_2021.zip' o 'march-2022.zip'.
    Restituisce il fine mese corrispondente, o None.
    """
    stem = path.stem.lower().replace("_", " ").replace("-", " ")
    tokens = stem.split()
    month_num = None
    year_num  = None
    for token in tokens:
        if token in MONTH_NAMES:
            month_num = MONTH_NAMES[token]
        elif token.isdigit() and len(token) == 4:
            year_num = int(token)
    if month_num and year_num:
        return _last_day(year_num, month_num)
    return None


def _parse_yyyymmdd_zip(path: Path) -> date | None:
    """
    Cerca una sequenza di 8 cifre che formi una data valida YYYYMMDD.
    Itera su tutti i match per gestire nomi con UUID prefisso
    (es. 1781298998603_EIOPA_RFR_20260531.zip).
    """
    for m in re.finditer(r"\d{8}", path.stem):
        try:
            d = datetime.strptime(m.group(), "%Y%m%d").date()
            if 2000 <= d.year <= 2100:
                return d
        except ValueError:
            pass
    return None


def _build_zip_index() -> dict[date, Path]:
    """
    Scansiona ZIP_FOLDER e restituisce un dict {fine_mese: path}
    accettando entrambi i formati di nome.
    """
    index: dict[date, Path] = {}
    folder = Path(ZIP_FOLDER)
    if not folder.exists():
        return index
    for p in folder.glob("*.zip"):
        ref = _parse_yyyymmdd_zip(p) or _parse_month_name_zip(p)
        if ref:
            index[ref] = p
    return index


def find_local_zip(ref_date: date, index: dict[date, Path] | None = None) -> Path | None:
    """
    Cerca il zip nella cartella ad ogni chiamata (glob diretto),
    in modo da vedere anche file aggiunti dopo l'avvio dello script.
    L'index opzionale viene ignorato — si preferisce sempre il glob fresco.
    """
    date_str = ref_date.strftime("%Y%m%d")
    folder = Path(ZIP_FOLDER)
    if not folder.exists():
        return None
    # Prima cerca per data numerica YYYYMMDD nel nome
    matches = list(folder.glob(f"*{date_str}*.zip"))
    if matches:
        return matches[0]
    # Poi cerca per nome del mese in inglese/italiano
    for p in folder.glob("*.zip"):
        if _parse_month_name_zip(p) == ref_date:
            return p
    return None


def download_zip_auto(ref_date: date) -> bytes | None:
    """Tenta download automatico e salva in eiopa_zips/."""
    date_str = ref_date.strftime("%Y%m%d")
    for template in AUTO_DOWNLOAD_TEMPLATES:
        try:
            r = requests.get(template.format(date=date_str), headers=HEADERS, timeout=30)
            if r.status_code == 200:
                out = Path(ZIP_FOLDER) / f"EIOPA_RFR_{date_str}.zip"
                out.write_bytes(r.content)
                return r.content
        except Exception:
            pass
    return None


def get_zip_bytes(ref_date: date, index: dict[date, Path] | None = None) -> tuple[bytes | None, str]:
    local = find_local_zip(ref_date)
    if local:
        return local.read_bytes(), f"locale ({local.name})"
    remote = download_zip_auto(ref_date)
    if remote:
        return remote, "scaricato automaticamente"
    return None, "non trovato"


def find_sheet(sheet_names: list) -> str | None:
    for s in sheet_names:
        if any(alias in s.lower().replace("-", "_") for alias in SHEET_ALIASES):
            return s
    return None


def parse_sheet(excel_bytes: bytes, sheet_name: str) -> list | None:
    # Metodo 1: formato nuovo (header riga 0, index col 0)
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        df = xl.parse(sheet_name, index_col=0, header=0)
        euro_col = next((c for c in df.columns if str(c).strip().lower() == "euro"), None)
        if euro_col:
            num_idx = pd.to_numeric(df.index, errors="coerce")
            df2 = df[num_idx.notna()].copy()
            df2.index = num_idx[num_idx.notna()]
            df2 = df2.sort_index().loc[1:20]
            values = df2[euro_col].dropna().tolist()
            if len(values) == 20:
                return [float(v) for v in values]
    except Exception:
        pass

    # Metodo 2: formato vecchio (header dinamico con metadata)
    try:
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        df_raw = xl.parse(sheet_name, header=None)
        header_row_idx = None
        for i, row in df_raw.iterrows():
            if any(str(v).strip().lower() == "euro" for v in row):
                header_row_idx = i
                break
        if header_row_idx is None:
            return None
        sub = df_raw.iloc[header_row_idx:].copy()
        sub.columns = sub.iloc[0]
        sub = sub.iloc[1:].set_index(sub.columns[1])
        numeric_mask = pd.to_numeric(sub.index, errors="coerce").notna()
        sub = sub[numeric_mask].copy()
        sub.index = pd.to_numeric(sub.index)
        sub = sub.sort_index().loc[1:20]
        euro_col = next((c for c in sub.columns if str(c).strip().lower() == "euro"), None)
        if euro_col is None:
            return None
        values = sub[euro_col].dropna().tolist()
        if len(values) == 20:
            return [float(v) for v in values]
    except Exception:
        pass
    return None


def extract_curve(zip_bytes: bytes) -> list | None:
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            excel_name = next(
                (n for n in z.namelist() if "Term_Structures" in n and n.endswith(".xlsx")),
                next((n for n in z.namelist() if n.endswith(".xlsx")), None),
            )
            if excel_name is None:
                return None
            excel_bytes = z.read(excel_name)
        xl = pd.ExcelFile(io.BytesIO(excel_bytes))
        sheet_name = find_sheet(xl.sheet_names)
        if sheet_name is None:
            return None
        return parse_sheet(excel_bytes, sheet_name)
    except Exception:
        return None


def build_xlsx(records: list, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EIOPA RFR EUR spot no VA"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", start_color="1F4E79")
    data_font   = Font(name="Arial", size=10)
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["TIME_PERIOD"] + [f"{t}Y" for t in TENORS]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, (ref_date, values) in enumerate(records, 2):
        dc = ws.cell(row=r, column=1, value=ref_date.strftime("%d/%m/%Y"))
        dc.font = data_font
        dc.alignment = Alignment(horizontal="center")
        dc.border = border
        for c, v in enumerate(values, 2):
            cell = ws.cell(row=r, column=c, value=round(v * 100, 9))
            cell.font = data_font
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
            cell.number_format = "#,##0.000000000"

    ws.column_dimensions["A"].width = 15
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B2"
    wb.save(output_path)
    print(f"\n✓ Salvato: {output_path}  ({len(records)} curve)")


def check_completeness(ref_dates: list, records: list) -> None:
    """
    Verifica che non ci siano buchi nella serie temporale prodotta.
    Stampa un riepilogo con i mesi mancanti tra il più vecchio e il più recente.
    """
    if not records:
        return

    produced = {d for d, _ in records}
    first    = min(produced)
    last     = max(produced)

    # Tutti i fine mese nell'intervallo coperto
    expected = {d.date() for d in pd.date_range(first, last, freq="ME")}
    missing  = sorted(expected - produced)

    print("\n── Controllo completezza ─────────────────────────────────")
    print(f"  Periodo coperto:  {first} → {last}")
    print(f"  Mesi attesi:      {len(expected)}")
    print(f"  Mesi prodotti:    {len(produced)}")

    if missing:
        print(f"  Mesi MANCANTI ({len(missing)}):  ← scarica e aggiungi in '{ZIP_FOLDER}/'")
        for d in missing:
            print(f"    {d}  →  EIOPA_RFR_{d.strftime('%Y%m%d')}.zip")
    else:
        print("  ✓ Serie completa, nessun mese mancante")
    print("──────────────────────────────────────────────────────────")


def print_missing(ref_dates: list, index: dict) -> None:
    missing = [d for d in ref_dates if d not in index]
    if not missing:
        print(f"✓ Tutti i {len(ref_dates)} zip sono presenti in '{ZIP_FOLDER}/'")
        return
    print(f"Zip mancanti ({len(missing)}) — scarica da EIOPA e metti in '{ZIP_FOLDER}/':")
    for d in missing:
        print(f"  EIOPA_RFR_{d.strftime('%Y%m%d')}.zip  ({d.strftime('%B %Y')})")


def diagnose(ref_date_str: str) -> None:
    ref_date = datetime.strptime(ref_date_str, "%Y-%m-%d").date()
    index = _build_zip_index()
    print(f"\n=== DIAGNOSI per {ref_date} ===")
    local = find_local_zip(ref_date)
    print(f"File locale: {local or f'non trovato in {ZIP_FOLDER}/'}")
    zip_bytes, source = get_zip_bytes(ref_date)
    if zip_bytes is None:
        print("Nessun zip disponibile.")
        return
    print(f"Zip: {len(zip_bytes)} bytes  (fonte: {source})")
    result = extract_curve(zip_bytes)
    if result:
        print(f"✓ Parsing OK")
        print(f"  1-5Y:   {[round(v*100, 4) for v in result[:5]]}")
        print(f"  16-20Y: {[round(v*100, 4) for v in result[15:]]}")
    else:
        print("✗ Parsing fallito")


def main() -> None:
    Path(ZIP_FOLDER).mkdir(exist_ok=True)

    all_ends  = pd.date_range(start=START_DATE, end=END_DATE, freq="ME")
    ref_dates = [d.date() for d in all_ends]

    # Costruisce l'indice una volta sola
    index = _build_zip_index()

    local_count = sum(1 for d in ref_dates if d in index)
    print(f"Date totali:  {len(ref_dates)}  ({ref_dates[0]} → {ref_dates[-1]})")
    print(f"Zip locali:   {local_count}  (riconosciuti in '{ZIP_FOLDER}/')")
    print(f"Da scaricare: {len(ref_dates) - local_count}\n")

    records = []
    for i, ref_date in enumerate(ref_dates, 1):
        print(f"[{i:3d}/{len(ref_dates)}] {ref_date} ...", end=" ", flush=True)
        zip_bytes, source = get_zip_bytes(ref_date)
        if zip_bytes is None:
            print(f"[SKIP] → metti il zip in '{ZIP_FOLDER}/'")
            continue
        result = extract_curve(zip_bytes)
        if result is None:
            print(f"[WARN] zip ok ({source}) ma parsing fallito")
            continue
        records.append((ref_date, result))
        print(f"OK ({source})")
        time.sleep(SLEEP_SEC)

    print(f"\nCurve prodotte: {len(records)}  |  Saltate: {len(ref_dates) - len(records)}")

    # Controllo completezza
    check_completeness(ref_dates, records)

    if records:
        build_xlsx(records, OUTPUT_FILE)


if __name__ == "__main__":
    if len(sys.argv) == 3 and sys.argv[1] == "diagnose":
        diagnose(sys.argv[2])
    elif len(sys.argv) == 2 and sys.argv[1] == "missing":
        Path(ZIP_FOLDER).mkdir(exist_ok=True)
        all_ends = pd.date_range(start=START_DATE, end=END_DATE, freq="ME")
        index    = _build_zip_index()
        print_missing([d.date() for d in all_ends], index)
    else:
        main()
